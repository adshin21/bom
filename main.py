from openpyxl import load_workbook, Workbook
from Solution import Solution


def parse_file():
    wb = load_workbook(FILE_LOCATION)
    sheet = wb['Source']

    row = sheet.max_row
    col = sheet.max_column

    parsed_data = {}
    data_info = {}
    for i in range(1, row):
        bad_row = False
        key = None
        values = []
        for j in range(col):
            val = sheet.cell(row=i + 1, column=j + 1).value

            if not val:
                bad_row = True
                break

            if j != 3:
                val = val.strip().strip('.')
            if j == 0:
                key = val
            else:
                if j == 1:
                    values.append(int(val))
                else:
                    values.append(val)

        if not bad_row and key:
            if key in parsed_data:
                parsed_data[key].append(values)
            else:
                parsed_data[key] = [values]
            data_info[values[1]] = (values[2], values[3])
    return parsed_data, data_info


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    FILE_LOCATION = 'BOM.xlsx'
    parsed_input_value, data_info = parse_file()

    workbook = Workbook()
    cls = Solution(parsed_input_value, data_info, workbook)
    cls.solve()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/

'''
Fan ::  ['1', 'Motor', 1, 'PC'] 
        ['2', 'Wires', 20, 'm'] 
        ['2', 'Plates', 2, 'PC'] 
        ['1', 'Blades', 3, 'PC'] 
        ['1', 'Screws', 10, 'PC']

Toy ::  ['1', 'Plastic sheet', 1, 'SqM'] 
        ['1', 'Metal Tools', 3, 'Pc'] 
        ['2', 'Iron chips', 10, 'Pc'] 
        ['2', 'Copper granule', 0.5, 'Kg'] 
        ['3', 'Copper Dust', 1, 'Kg'] 
        ['1', 'Packaging Box', 1, 'Pc']
'''