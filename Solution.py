from collections import defaultdict
from queue import Queue

from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.styles.differential import DifferentialStyle

class Solution:

    def __init__(self, raw_data, mapping, workbook):
        self.raw_data = raw_data
        self.mapping = mapping
        self.work_book = workbook

    def populateparents(self):
        raw_data = self.raw_data
        parents = {}
        for good in raw_data:
            stack = [(good, 0)]
            parents[good] = []
            for raw_material in raw_data[good]:
                current_level = raw_material[0]
                parent_level = stack[-1][1]
                # breakpoint()

                if current_level == parent_level + 1:
                    parents[good].append(stack[-1][0])
                    stack.append((raw_material[1], raw_material[0]))
                elif current_level <= parent_level:
                    while len(stack) and current_level <= parent_level:
                        stack.pop()
                        parent_level = stack[-1][1]

                    if len(stack):
                        parents[good].append(stack[-1][0])
                        stack.append((raw_material[1], raw_material[0]))
                    else:
                        parents[good].append(good)
                        stack.append((good, 0))
                else:
                    print('-------------------------')
                    print(current_level, parent_level)
                    print(raw_material)
                    print('-------------------------')
        return parents

    def generategraph(self, parents):
        graph = defaultdict(list)
        for key in self.raw_data:
            graph[key].append("root")
            graph["root"].append(key)

        for key in self.raw_data:
            for i, j in zip(self.raw_data[key], parents[key]):
                graph[i[1]].append(j)
                graph[j].append(i[1])
        return graph

    def createnewexcelsheet(self, data):

        ws = self.work_book.create_sheet(data["finished_good"])
        ws.title = data["finished_good"]
        dxf = DifferentialStyle(font=Font(bold=True), border=Border(), fill=PatternFill(bgColor="a1bdd5", fill_type="solid"))
        rule = Rule(type='cellIs', dxf=dxf)

        dxf1 = DifferentialStyle(fill=PatternFill(bgColor="FFFF00", fill_type="solid"))
        rule1 = Rule(type="cellIs", dxf=dxf1)
        ws.append(["Finished Good List"])
        ws.append(["#", "Item Description", "Quantity", "Unit"])
        ws.conditional_formatting.add("A2:D2", rule)
        ws.append([1, data["finished_good"], 1, "Pc"])
        ws.conditional_formatting.add("B3", rule1)
        ws.append(["End of FG"])
        ws.append([])
        ws.append(["Raw Material List"])
        ws.append(["#", "Item Description", "Quantity", "Unit"])
        ws.conditional_formatting.add("A7:D7", rule)

        for id, each in enumerate(data["raw_material"]):
            ws.append([id+1, each, self.mapping[each][0], self.mapping[each][1]])
            row = 7 + id
            start = "A" + str(row) + ':' + "D" + str(row)
            ws.conditional_formatting.add(start, rule1)
        ws.append(["End of RM"])

    def bfs(self, graph):
        q = Queue()
        q.put('root')
        lvl = 0
        vis = defaultdict(bool)
        while not q.empty():
            first = q.get()
            vis[first] = True
            lvl += 1
            finished_good = first
            raw_material = []
            for ele in graph[first]:
                if not vis[ele]:
                    q.put(ele)
                    raw_material.append(ele)

            if len(raw_material) and lvl > 1:
                self.createnewexcelsheet({
                    "finished_good": finished_good,
                    "raw_material": raw_material
                })

    def solve(self):
        parents = self.populateparents()
        graph = self.generategraph(parents)
        self.bfs(graph)
        self.work_book.save("Output.xlsx")
